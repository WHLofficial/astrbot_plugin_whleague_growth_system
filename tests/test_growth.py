"""pytest 单元测试：规则解析、经验计算、里程碑幂等、成长期推进、命令参数解析。"""

import asyncio
import json
import os
import tempfile

import pytest

from astrbot_plugin_whleague_growth_system.config.defaults import DEFAULT_CONFIG, validate_and_cast
from astrbot_plugin_whleague_growth_system.db.connection import DatabaseManager
from astrbot_plugin_whleague_growth_system.db.dao import GrowthDAO
from astrbot_plugin_whleague_growth_system.db.schema import init_schema
from astrbot_plugin_whleague_growth_system.services.growth_service import GrowthService
from astrbot_plugin_whleague_growth_system.services.export_service import (
    ExportService,
    _HEADERS,
)
from astrbot_plugin_whleague_growth_system.services.import_service import GrowthImportService, kind_from_name
from astrbot_plugin_whleague_growth_system.services.rule_parser import (
    RuleError,
    format_rule,
    normalize_rule,
    parse_rule_json,
    parse_rule_table,
)
from astrbot_plugin_whleague_growth_system.utils.messages import build_help, deny, usage
from astrbot_plugin_whleague_growth_system.utils.security import (
    parse_date,
    parse_num,
    sanitize_filename,
    sanitize_uid,
)


# ─── 规则解析 ─────────────────────────────────────────────

def _rule_json():
    return {
        "stats": {
            "goal": {"name": "进球", "xp": 10},
            "assist": {"name": "助攻", "xp": 5},
            "appearance": {"name": "出场", "xp": 2},
        },
        "milestones": [
            {"stat": "goal", "period": "period", "threshold": 10, "xp": 50},
            {"stat": "goal", "period": "career", "threshold": 100, "xp": 1000},
        ],
        "level_xp": 100,
    }


def test_parse_rule_json_ok():
    rule = parse_rule_json(json.dumps(_rule_json()), 100)
    assert rule["stats"]["goal"]["xp"] == 10
    assert rule["milestones"][0]["period"] == "period"
    assert rule["level_xp"] == 100


def test_parse_rule_json_default_level_xp():
    data = _rule_json()
    del data["level_xp"]
    rule = parse_rule_json(json.dumps(data), 50)
    assert rule["level_xp"] == 50


def test_parse_rule_json_stat_shorthand():
    data = {"stats": {"goal": 10, "assist": {"name": "助攻", "xp": 5}}, "milestones": []}
    rule = normalize_rule(data, 100)
    assert rule["stats"]["goal"] == {"name": "goal", "xp": 10}


def test_parse_rule_json_invalid():
    with pytest.raises(RuleError):
        parse_rule_json("not json", 100)
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps({"stats": {}}), 100)
    bad = _rule_json()
    bad["milestones"] = [{"stat": "nope", "period": "period", "threshold": 1, "xp": 1}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    bad2 = _rule_json()
    bad2["milestones"] = [{"stat": "goal", "period": "yearly", "threshold": 1, "xp": 1}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad2), 100)


# ─── bands（区间经验）─────────────────────────────────────

def _bands_rule():
    return {
        "stats": {
            "goal": {"name": "进球", "xp": 10},
            "rating": {
                "name": "评分",
                "bands": [
                    {"min": 4.0, "max": 6.0, "xp": 5},
                    {"min": 6.0, "max": 8.0, "xp": 10},
                    {"min": 8.0, "xp": 20},
                ],
            },
        },
        "milestones": [],
        "level_xp": 100,
    }


def test_parse_rule_json_bands_ok():
    rule = parse_rule_json(json.dumps(_bands_rule()), 100)
    bands = rule["stats"]["rating"]["bands"]
    assert bands == [
        {"min": 4.0, "max": 6.0, "xp": 5},
        {"min": 6.0, "max": 8.0, "xp": 10},
        {"min": 8.0, "xp": 20},
    ]
    assert "xp" not in rule["stats"]["rating"]


def test_parse_rule_json_bands_invalid():
    base = _bands_rule()
    # 重叠
    bad = json.loads(json.dumps(base))
    bad["stats"]["rating"]["bands"] = [
        {"min": 4.0, "max": 8.0, "xp": 5},
        {"min": 6.0, "max": 8.0, "xp": 10},
    ]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # 开放上界非末段
    bad = json.loads(json.dumps(base))
    bad["stats"]["rating"]["bands"] = [
        {"min": 4.0, "xp": 5},
        {"min": 6.0, "max": 8.0, "xp": 10},
    ]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # 下限为负
    bad = json.loads(json.dumps(base))
    bad["stats"]["rating"]["bands"] = [{"min": -1, "max": 6.0, "xp": 5}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # xp 与 bands 并存
    bad = json.loads(json.dumps(base))
    bad["stats"]["rating"]["xp"] = 10
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # max <= min
    bad = json.loads(json.dumps(base))
    bad["stats"]["rating"]["bands"] = [{"min": 6.0, "max": 4.0, "xp": 5}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)


def test_bands_cannot_be_milestone_stat():
    base = _bands_rule()
    bad = json.loads(json.dumps(base))
    bad["milestones"] = [{"stat": "rating", "period": "period", "threshold": 50, "xp": 100}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    bad2 = json.loads(json.dumps(base))
    bad2["milestones"] = [{"stat": "rating", "period": "period", "step": 10, "xp": 100}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad2), 100)


def test_bands_xp_calculation():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_bands_rule()), 100)
            await service.save_rule(rule, "bands", "admin")

            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")

            await service._db.execute_transaction(_tx)
            # 5.0 → [4,6) 得 5；线性项不受影响
            r = await service.record_match("p01", "2026-08-01", "", {"rating": 5.0, "goal": 1}, "admin")
            assert r["stat_xp"] == 5 + 10
            # 7.5 → [6,8) 得 10
            r = await service.record_match("p01", "2026-08-02", "", {"rating": 7.5}, "admin")
            assert r["stat_xp"] == 10
            # 9.0 / 10.0 / 11.0 → [8,∞) 开放段得 20
            r = await service.record_match("p01", "2026-08-03", "", {"rating": 9.0}, "admin")
            assert r["stat_xp"] == 20
            r = await service.record_match("p01", "2026-08-03b", "", {"rating": 10.0}, "admin")
            assert r["stat_xp"] == 20
            # 未命中（低于首段下限 3.5）得 0，不报错
            r = await service.record_match("p01", "2026-08-04", "", {"rating": 3.5}, "admin")
            assert r["stat_xp"] == 0
            # 边界值：4.0 命中首段，6.0 命中次段（左闭右开）
            r = await service.record_match("p01", "2026-08-06", "", {"rating": 4.0}, "admin")
            assert r["stat_xp"] == 5
            r = await service.record_match("p01", "2026-08-07", "", {"rating": 6.0}, "admin")
            assert r["stat_xp"] == 10
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


# ─── repeat（每累计 n 次奖励，可重复）─────────────────────

def _repeat_rule():
    return {
        "stats": {
            "appearance": {"name": "出场", "xp": 2},
            "goal": {"name": "进球", "xp": 10},
        },
        "milestones": [
            {"stat": "appearance", "period": "period", "step": 10, "xp": 50},
        ],
        "level_xp": 100,
    }


def test_parse_rule_json_repeat():
    rule = parse_rule_json(json.dumps(_repeat_rule()), 100)
    assert rule["milestones"] == [
        {"stat": "appearance", "period": "period", "step": 10, "xp": 50}
    ]
    # step 与 threshold 并存报错
    bad = json.loads(json.dumps(_repeat_rule()))
    bad["milestones"] = [
        {"stat": "appearance", "period": "period", "step": 10, "threshold": 5, "xp": 50}
    ]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # step 非正报错
    bad = json.loads(json.dumps(_repeat_rule()))
    bad["milestones"] = [{"stat": "appearance", "period": "period", "step": 0, "xp": 50}]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)


def test_repeat_rewards_multiple_times():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_repeat_rule()), 100)
            await service.save_rule(rule, "repeat", "admin")

            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")

            await service._db.execute_transaction(_tx)
            # 每场 5 次出场（数据经验 10/场）
            r1 = await service.record_match("p01", "2026-08-01", "", {"appearance": 5}, "admin")
            assert r1["bonus_xp"] == 0
            # 累计 10 → 触发一次重复奖励 +50
            r2 = await service.record_match("p01", "2026-08-08", "", {"appearance": 5}, "admin")
            assert r2["bonus_xp"] == 50
            assert len(r2["awarded"]) == 1 and r2["awarded"][0]["count"] == 1
            p = await dao.get_player("p01")
            assert p["xp"] == 10 + 10 + 50
            # 再 5 次 → 累计 15，仍只有 1 档
            r3 = await service.record_match("p01", "2026-08-15", "", {"appearance": 5}, "admin")
            assert r3["bonus_xp"] == 0
            p = await dao.get_player("p01")
            assert p["xp"] == 20 + 50 + 10
            # 再 5 次 → 累计 20 → 第二档 +50
            r4 = await service.record_match("p01", "2026-08-22", "", {"appearance": 5}, "admin")
            assert r4["bonus_xp"] == 50
            assert r4["awarded"][0]["count"] == 1
            p = await dao.get_player("p01")
            assert p["xp"] == 30 + 100 + 10
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_repeat_resets_on_period_advance():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_repeat_rule()), 100)
            await service.save_rule(rule, "repeat", "admin")

            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")

            await service._db.execute_transaction(_tx)
            await service.record_match("p01", "2026-08-01", "", {"appearance": 10}, "admin")
            assert (await dao.get_player("p01"))["xp"] == 20 + 50
            # 推进成长期后重新累计
            await service.advance_period("成长期2", True)
            r = await service.record_match("p01", "2026-09-01", "", {"appearance": 10}, "admin")
            assert r["bonus_xp"] == 50, "新成长期应重新触发"
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


# ─── 表格解析：band / repeat 行 ───────────────────────────

def test_parse_rule_table_band_and_repeat():
    rows = [
        ["type", "stat", "name", "xp", "period", "threshold", "min", "max"],
        ["stat", "goal", "进球", "10", "", "", "", ""],
        ["stat", "appearance", "出场", "2", "", "", "", ""],
        ["band", "rating", "评分", "5", "", "", "4.0", "6.0"],
        ["band", "rating", "", "10", "", "", "6.0", "8.0"],
        ["band", "rating", "", "20", "", "", "8.0", ""],
        ["repeat", "appearance", "", "50", "period", "10", "", ""],
        ["level", "", "", "100", "", "", "", ""],
    ]
    rule = parse_rule_table(rows, DEFAULT_CONFIG, 100)
    assert rule["stats"]["goal"] == {"name": "进球", "xp": 10}
    assert rule["stats"]["rating"]["name"] == "评分"
    assert rule["stats"]["rating"]["bands"] == [
        {"min": 4.0, "max": 6.0, "xp": 5},
        {"min": 6.0, "max": 8.0, "xp": 10},
        {"min": 8.0, "xp": 20},
    ]
    assert rule["milestones"] == [
        {"stat": "appearance", "period": "period", "step": 10, "xp": 50}
    ]
    assert rule["level_xp"] == 100


def test_parse_rule_table_linear_band_conflict():
    rows = [
        ["type", "stat", "name", "xp", "period", "threshold", "min", "max"],
        ["stat", "rating", "评分", "10", "", "", "", ""],
        ["band", "rating", "", "5", "", "", "4.0", "6.0"],
    ]
    with pytest.raises(RuleError):
        parse_rule_table(rows, DEFAULT_CONFIG, 100)


def test_parse_rule_table_band_no_type_col():
    # 无类型列布局（列位左移一列）：stat/name/xp/period/threshold/min/max
    rows = [
        ["goal", "进球", 10],
        ["rating", "评分", 5, "", "", 4.0, 6.0],
        ["rating", "", 10, "", "", 6.0, 8.0],
    ]
    rule = parse_rule_table(rows, DEFAULT_CONFIG, 100)
    assert rule["stats"]["goal"]["xp"] == 10
    assert rule["stats"]["rating"]["bands"][0] == {"min": 4.0, "max": 6.0, "xp": 5}


def test_format_rule_bands_and_repeat():
    rule = parse_rule_json(json.dumps(_bands_rule()), 100)
    text = format_rule(rule)
    assert "评分" in text and "[4~6)+5" in text and "[8~)+20" in text
    assert "未命中区间得 0" in text
    rule2 = parse_rule_json(json.dumps(_repeat_rule()), 100)
    text2 = format_rule(rule2)
    assert "每累计 10 次 → 奖励 50 经验（可重复）" in text2


def test_parse_rule_table_csv():
    rows = [
        ["type", "stat", "name", "xp", "period", "threshold"],
        ["stat", "goal", "进球", "10", "", ""],
        ["stat", "assist", "助攻", "5", "", ""],
        ["milestone", "goal", "", "50", "period", "10"],
        ["level", "", "", "100", "", ""],
    ]
    rule = parse_rule_table(rows, DEFAULT_CONFIG, 100)
    assert rule["stats"]["goal"]["name"] == "进球"
    assert rule["stats"]["goal"]["xp"] == 10
    assert rule["milestones"] == [
        {"stat": "goal", "period": "period", "threshold": 10.0, "xp": 50}
    ]
    assert rule["level_xp"] == 100


def test_parse_rule_table_type_inference():
    # 无类型列布局: [stat, name, xp, period, threshold]
    rows = [
        ["goal", "进球", 10],
        ["assist", "助攻", 5],
        ["goal", "", 50, "period", 10],
    ]
    rule = parse_rule_table(rows, DEFAULT_CONFIG, 100)
    assert "goal" in rule["stats"]
    assert rule["stats"]["goal"]["xp"] == 10
    assert rule["milestones"][0] == {
        "stat": "goal", "period": "period", "threshold": 10.0, "xp": 50
    }


# ─── 文件类型嗅探 ─────────────────────────────────────────

def test_kind_from_name():
    assert kind_from_name("规则_进球.json") == "rule"
    assert kind_from_name("球员_名单.csv") == "players"
    assert kind_from_name("比赛_2026.csv") == "matches"
    assert kind_from_name("report.csv") is None


def test_kind_from_name_instance_method():
    """回归：群文件捕获/命令按实例方法调用 kind_from_name，缺失会 AttributeError 崩溃。"""
    service, dao, imp, tmp, env = _make_env()
    try:
        assert imp.kind_from_name("规则_a.json") == "rule"
        assert imp.kind_from_name("球员_a.csv") == "players"
        assert imp.kind_from_name("比赛_a.csv") == "matches"
        assert imp.kind_from_name("report.csv") is None
    finally:
        asyncio.run(env["db"].close())


# ─── 工具函数 ─────────────────────────────────────────────

def test_parse_date_variants():
    assert parse_date("2026-08-14") == "2026-08-14"
    assert parse_date("2026/8/14") == "2026-08-14"
    assert parse_date("20260814") == "2026-08-14"
    with pytest.raises(ValueError):
        parse_date("14-08-2026")


def test_parse_date_invalid():
    # 非法日期必须拒绝（修复3 回归）
    for bad in ("2026-13-40", "2026-02-30", "2026-00-01", "2026-13-01", "2026-00-00"):
        with pytest.raises(ValueError):
            parse_date(bad)


def test_parse_num():
    assert parse_num("2") == 2.0
    assert parse_num("1.5") == 1.5
    with pytest.raises(ValueError):
        parse_num("-1")
    with pytest.raises(ValueError):
        parse_num("abc")


def test_sanitize():
    assert sanitize_uid("  p01\n") == "p01"
    assert sanitize_filename("../恶意/文件.csv") == ".._恶意_文件.csv"


def test_validate_and_cast():
    assert validate_and_cast("default_level_xp", "150") == 150
    with pytest.raises(ValueError):
        validate_and_cast("default_level_xp", "0")
    assert validate_and_cast("advance_default_carryover", "false") is False
    assert validate_and_cast("group_whitelist", "111,222") == ["111", "222"]


# ─── 经验值支持最多 1 位小数 ──────────────────────────────

def _decimal_rule():
    return {
        "stats": {
            "goal": {"name": "进球", "xp": 2.5},
            "rating": {
                "name": "评分",
                "bands": [{"min": 4.0, "max": 6.0, "xp": 1.5}, {"min": 6.0, "xp": 3.5}],
            },
        },
        "milestones": [
            {"stat": "goal", "period": "period", "threshold": 10, "xp": 12.5},
            {"stat": "goal", "period": "period", "step": 5, "xp": 7.5},
        ],
        "level_xp": 12.5,
    }


def test_parse_decimal_xp():
    rule = parse_rule_json(json.dumps(_decimal_rule()), 100)
    assert rule["stats"]["goal"]["xp"] == 2.5
    assert rule["stats"]["rating"]["bands"][0]["xp"] == 1.5
    assert rule["milestones"][0]["xp"] == 12.5
    assert rule["milestones"][1]["xp"] == 7.5
    assert rule["level_xp"] == 12.5
    # 2 位小数报错
    bad = json.loads(json.dumps(_decimal_rule()))
    bad["stats"]["goal"]["xp"] = 0.25
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # 非正报错
    bad = json.loads(json.dumps(_decimal_rule()))
    bad["stats"]["goal"]["xp"] = 0
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)
    # step 仍要求整数
    bad = json.loads(json.dumps(_decimal_rule()))
    bad["milestones"][1]["step"] = 2.5
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(bad), 100)


def test_decimal_xp_calculation():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_decimal_rule()), 100)
            await service.save_rule(rule, "dec", "admin")

            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")

            await service._db.execute_transaction(_tx)
            # 线性：3 * 2.5 = 7.5；bands：5.0 命中 [4,6) 得 1.5
            r = await service.record_match(
                "p01", "2026-08-01", "", {"goal": 3, "rating": 5.0}, "admin"
            )
            assert r["stat_xp"] == 9.0  # 7.5 + 1.5
            p = await dao.get_player("p01")
            assert p["xp"] == 9.0 and p["xp_total"] == 9.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_decimal_overwrite_delta():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_decimal_rule()), 100)
            await service.save_rule(rule, "dec", "admin")

            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")

            await service._db.execute_transaction(_tx)
            await service.record_match("p01", "2026-08-01", "", {"goal": 3}, "admin")
            # 覆盖为 goal=1 → 2.5，delta = 2.5 - 7.5 = -5，xp 从 7.5 回到 2.5
            r = await service.record_match("p01", "2026-08-01", "", {"goal": 1}, "admin")
            assert r["stat_xp"] == 2.5
            p = await dao.get_player("p01")
            assert p["xp"] == 2.5
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_advance_decimal_settlement():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_decimal_rule()), 100)
            await service.save_rule(rule, "dec", "admin")

            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")

            await service._db.execute_transaction(_tx)
            # 直接构造小数 xp，绕开录入的里程碑干扰
            await service._db.execute(
                "UPDATE players SET xp=25.0, xp_total=25.0 WHERE player_uid='p01'"
            )
            # xp=25.0, level_xp=12.5 → 升 2 级，溢出 0（整数运算精确）
            r = await service.advance_period("成长期2", True)
            assert r["level_xp"] == 12.5
            p = await dao.get_player("p01")
            assert p["level"] == 3 and p["xp"] == 0.0
            # 带溢出：xp=12.5 → 升 1 级，溢出 0；xp=37.5 → 升 3 级溢出 0
            await service._db.execute(
                "UPDATE players SET xp=37.5, xp_total=62.5 WHERE player_uid='p01'"
            )
            await service.advance_period("成长期3", True)
            p = await dao.get_player("p01")
            assert p["level"] == 6 and p["xp"] == 0.0 and p["xp_total"] == 62.5
            # 溢出结转：xp=45.0 → 升 3 级（37.5/12.5=3）溢出 7.5
            await service._db.execute(
                "UPDATE players SET xp=45.0, xp_total=107.5 WHERE player_uid='p01'"
            )
            await service.advance_period("成长期4", True)
            p = await dao.get_player("p01")
            assert p["level"] == 9 and p["xp"] == 7.5 and p["xp_total"] == 107.5
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_fmt_xp_display():
    from astrbot_plugin_whleague_growth_system.utils.security import fmt_xp
    assert fmt_xp(10.0) == "10"
    assert fmt_xp(12.5) == "12.5"
    assert fmt_xp(0.0) == "0"
    assert fmt_xp(None) == "0"
    assert fmt_xp(7.5) == "7.5"


def test_format_rule_decimal_display():
    rule = parse_rule_json(json.dumps(_decimal_rule()), 100)
    text = format_rule(rule)
    assert "每单位 2.5 经验" in text
    assert "[4~6)+1.5" in text
    assert "奖励 12.5 经验" in text
    assert "每累计 5 次 → 奖励 7.5 经验（可重复）" in text
    assert "每级所需经验：12.5" in text


# ─── 统一反馈文案（utils/messages）─────────────────────────

def test_usage_format():
    assert usage("成长查询", "<球员ID>") == "用法: /成长查询 <球员ID>"
    assert usage("成长上报", "<球员ID> <日期> <数据项=值>...", "/成长上报 p01 2026-08-14 进球=2") == (
        "用法: /成长上报 <球员ID> <日期> <数据项=值>...\n例: /成长上报 p01 2026-08-14 进球=2"
    )


def test_deny_hints_admin_ids():
    assert deny().startswith("该命令需要管理员权限。")
    assert "admin_ids" in deny()


def test_build_help_requires_admin():
    player_text = build_help(False)
    assert "管理命令：" not in player_text
    assert "/成长上报" not in player_text
    assert "/成长规则" in player_text
    admin_text = build_help(True)
    assert "管理命令：" in admin_text
    assert "/成长设置" in admin_text
    assert "/成长查看配置" in admin_text


def test_error_hints_in_service_messages():
    # 服务层报错原文带修正引导
    with pytest.raises(ValueError) as e:
        parse_date("14/08/2026")
    assert "例: 2026-08-14" in str(e.value)
    with pytest.raises(ValueError) as e:
        parse_num("abc")
    assert "需为非负数字" in str(e.value)


# ─── 集成：经验 / 里程碑 / 推进（内存 SQLite）──────────────

def _make_env():
    """构造临时环境，返回 (service, dao, imp, tmp_dir)。"""
    tmp = tempfile.mkdtemp()
    db = DatabaseManager(os.path.join(tmp, "test.db"))
    # 环境初始化在调用方的事件循环内完成
    asyncio.run(db.init())
    asyncio.run(init_schema(db))
    dao = GrowthDAO(db)
    service = GrowthService(db, dao, DEFAULT_CONFIG.get)
    imp = GrowthImportService(db, dao, DEFAULT_CONFIG.get, service)
    env_holder = {"db": db}
    return service, dao, imp, tmp, env_holder


def _run_async(coro):
    return asyncio.run(coro)


async def _setup(service, dao):
    rule = parse_rule_json(json.dumps(_rule_json()), 100)
    await service.save_rule(rule, "test", "admin")

    async def _tx(conn):
        await dao.upsert_player(conn, "p01", "球员一", "A队", "admin")
        await dao.upsert_player(conn, "p02", "球员二", "B队", "admin")

    await service._db.execute_transaction(_tx)


def test_record_match_and_xp():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            r = await service.record_match("p01", "2026-08-01", "强队", {"goal": 2, "assist": 1}, "admin")
            assert r["stat_xp"] == 25  # 2*10 + 1*5
            assert r["total_xp"] == 25
            assert r["xp"] == 25
            assert r["xp_total"] == 25
            p = await dao.get_player("p01")
            assert p["xp"] == 25 and p["xp_total"] == 25
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_record_match_overwrite_same_day():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 1}, "admin")
            r = await service.record_match("p01", "2026-08-01", "", {"goal": 3}, "admin")
            # 覆盖：同日期同球员替换整条
            assert r["stat_xp"] == 30
            p = await dao.get_player("p01")
            assert p["xp"] == 30
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_milestone_award_and_idempotency():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            # 两场各 5 球，累计 10 达成成长期里程碑（奖励 50）
            r1 = await service.record_match("p01", "2026-08-01", "", {"goal": 5}, "admin")
            assert r1["bonus_xp"] == 0
            r2 = await service.record_match("p01", "2026-08-08", "", {"goal": 5}, "admin")
            assert r2["bonus_xp"] == 50
            assert len(r2["awarded"]) == 1
            p = await dao.get_player("p01")
            assert p["xp"] == 100 + 50  # 10*10 数据 + 50 奖励
            # 再录一场：累计超阈值但已颁发，不重复奖励
            r3 = await service.record_match("p01", "2026-08-15", "", {"goal": 1}, "admin")
            assert r3["bonus_xp"] == 0
            p = await dao.get_player("p01")
            assert p["xp"] == 100 + 50 + 10
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_career_milestone_uses_total():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            # 生涯阈值 100：分两成长期累计（推进后数据仍计入生涯）
            await service.record_match("p01", "2026-08-01", "", {"goal": 60}, "admin")
            await service.advance_period("成长期2", True)
            r = await service.record_match("p01", "2026-09-01", "", {"goal": 50}, "admin")
            # 生涯里程碑 +1000；新成长期累计 50 球也触发 period 里程碑 +50
            assert r["bonus_xp"] == 1050
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_advance_carryover_and_clear():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            # 25 球 → 数据经验 250 + 成长期里程碑(10球)奖励 50 → xp=300
            # 每级 100 → 升 3 级，溢出 0
            r = await service.advance_period("成长期2", True)
            assert r["opened_no"] == 2
            p = await dao.get_player("p01")
            assert p["level"] == 4
            assert p["xp"] == 0  # 溢出 0
            assert p["xp_total"] == 300  # 生涯经验不动
            # 清零分支
            await service.record_match("p01", "2026-09-01", "", {"goal": 5}, "admin")
            p = await dao.get_player("p01")
            assert p["xp"] == 50
            r2 = await service.advance_period("成长期3", False)
            assert r2["opened_no"] == 3
            p = await dao.get_player("p01")
            assert p["level"] == 4  # 50//100=0 不升级
            assert p["xp"] == 0  # 清零
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_advance_no_rule_uses_default():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "球员一", "", "admin")

            await service._db.execute_transaction(_tx)
            # 无规则时按 default_level_xp=100
            await service._db.execute(
                "UPDATE players SET xp=250 WHERE player_uid='p01'"
            )
            r = await service.advance_period("成长期2", True)
            assert r["level_xp"] == 100
            p = await dao.get_player("p01")
            assert p["level"] == 3 and p["xp"] == 50
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_record_unknown_player_and_stat():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            with pytest.raises(ValueError):
                await service.record_match("nobody", "2026-08-01", "", {"goal": 1}, "admin")
            with pytest.raises(ValueError):
                await service.record_match("p01", "2026-08-01", "", {"unknown": 1}, "admin")
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_rule_import_roundtrip():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_rule_json()), 100)
            await service.save_rule(rule, "x.json", "admin")
            loaded = await service.get_rule()
            assert loaded["stats"]["goal"]["xp"] == 10
            assert loaded["milestones"][1]["threshold"] == 100
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_overwrite_keeps_period():
    """跨成长期覆盖旧比赛必须保留原成长期，防止篡改历史统计（修复1 回归）。"""
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-07-01", "", {"goal": 1}, "admin")
            row = await dao._db.fetchone(
                "SELECT period_no FROM appearances a JOIN matches m ON m.id=a.match_id "
                "WHERE m.match_date='2026-07-01' AND a.player_uid='p01'"
            )
            assert row["period_no"] == 1
            await service.advance_period("成长期2", True)
            # 推进后覆盖成长期1的同一场比赛
            await service.record_match("p01", "2026-07-01", "", {"goal": 2}, "admin")
            row = await dao._db.fetchone(
                "SELECT period_no FROM appearances a JOIN matches m ON m.id=a.match_id "
                "WHERE m.match_date='2026-07-01' AND a.player_uid='p01'"
            )
            assert row["period_no"] == 1, "覆盖不得改写原成长期"
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_advance_concurrent_race():
    """并发推进必须只有一个当前成长期、球员只结算一次（修复2 回归）。"""
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            async def _tx(conn):
                await dao.upsert_player(conn, "p01", "P1", "", "admin")
                await dao.upsert_player(conn, "p02", "P2", "", "admin")
            await service._db.execute_transaction(_tx)
            await service._db.execute(
                "UPDATE players SET xp=150 WHERE player_uid IN ('p01','p02')"
            )
            results = await asyncio.gather(
                service.advance_period("期A", True),
                service.advance_period("期B", True),
                return_exceptions=True,
            )
            ok = [r for r in results if not isinstance(r, Exception)]
            errs = [r for r in results if isinstance(r, Exception)]
            assert len(ok) == 1, f"应恰好一次推进成功, 实际 {len(ok)}"
            assert errs, "并发第二个推进应失败"
            cur = await dao._db.fetchone(
                "SELECT COUNT(*) AS c FROM growth_periods WHERE is_current=1"
            )
            assert cur["c"] == 1, "只允许一个当前成长期"
            p = await dao.get_player("p01")
            assert p["level"] == 2 and p["xp"] == 50, "球员只应结算一次"
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_matches_empty_stats_rejected():
    """比赛文件无表头/无匹配列时不得生成 0 经验记录（修复4 回归）。"""
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            rule = parse_rule_json(json.dumps(_rule_json()), 100)
            file_path = os.path.join(tmp, "比赛_空表.csv")
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("2026-08-01,p01,3\n2026-08-02,p02,1\n")
            entries, errors, skipped = imp.parse_matches_file(file_path, rule)
            assert entries == [], "无有效数据列时不得生成记录"
            assert len(errors) == 2, "两行都应记为错误"
            # 带表头的正常文件
            file2 = os.path.join(tmp, "比赛_正常.csv")
            with open(file2, "w", encoding="utf-8") as f:
                f.write("日期,球员ID,进球,助攻\n2026-08-01,p01,2,1\n")
            entries2, errors2, skipped2 = imp.parse_matches_file(file2, rule)
            assert len(entries2) == 1 and entries2[0]["stats"] == {"goal": 2.0, "assist": 1.0}
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_config_bool_false_parsed():
    """advance_default_carryover 为字符串 'false' 时必须走清零分支（修复6 回归）。"""
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            from astrbot_plugin_whleague_growth_system.handlers.admin import _as_bool
            assert _as_bool(False, True) is False
            assert _as_bool("false", True) is False
            assert _as_bool("0", True) is False
            assert _as_bool("no", True) is False
            assert _as_bool("true", False) is True
            assert _as_bool("garbage", True) is True  # 非法值回退默认
            assert _as_bool(None, True) is True
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


# ─── 成长期结果快照 ──────────────────────────────────────

def test_period_snapshot_recorded_on_advance():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            # 规则含成长期里程碑（goal 累计10→+50）：
            # p01 录 25 球 → 数据250+里程碑50=300 → 升3级溢出0
            # p02 录 6 球 → 60 → 升0级溢出60
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            await service.record_match("p02", "2026-08-01", "", {"goal": 6}, "admin")
            r = await service.advance_period("成长期2", True)
            assert r["opened_no"] == 2
            # 快照存在且字段正确
            snaps = await dao.list_period_snapshots(1)
            by_uid = {s["player_uid"]: s for s in snaps}
            assert set(by_uid) == {"p01", "p02"}
            s1 = by_uid["p01"]
            assert s1["level_end"] == 4 and s1["level_gained"] == 3
            assert s1["xp_period"] == 300.0 and s1["xp_carryover"] == 0.0
            assert s1["player_name"] == "球员一"
            s2 = by_uid["p02"]
            assert s2["level_end"] == 1 and s2["level_gained"] == 0
            assert s2["xp_period"] == 60.0 and s2["xp_carryover"] == 60.0
            # 摘要聚合
            summary = await dao.summarize_period(1)
            assert summary["player_count"] == 2
            assert summary["upgraded_count"] == 1
            assert summary["xp_total"] == 360.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_period_snapshot_clear_strategy_carryover_zero():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            # 清零策略：结转记 0；p01 xp=300（含里程碑+50）
            await service.advance_period("成长期2", False)
            snaps = await dao.list_period_snapshots(1)
            s = [x for x in snaps if x["player_uid"] == "p01"][0]
            assert s["level_end"] == 4 and s["level_gained"] == 3
            assert s["xp_period"] == 300.0 and s["xp_carryover"] == 0.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_period_status_and_result():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            await service.advance_period("成长期2", True)
            # 仅 p01 有数据（300→升3溢出0），p02 未录（0）；当前期总经验 = 0
            st = await service.period_status()
            assert st["current"]["period_no"] == 2
            assert st["current_xp"] == 0.0
            assert len(st["summaries"]) == 1
            assert st["summaries"][0]["player_count"] == 2
            assert st["summaries"][0]["upgraded_count"] == 1
            assert st["summaries"][0]["xp_total"] == 300.0
            # 期结果明细
            res = await service.period_result(1)
            assert res is not None and res["period"]["period_no"] == 1
            assert len(res["rows"]) == 2
            # 不存在的期号 / 无快照期
            assert await service.period_result(99) is None
            assert await service.period_result(2) is None  # 当期尚无快照
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


# ─── 成长数据导出 / 预览 ─────────────────────────────────

def test_export_period_rows():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            # p01 25 球 → 250+里程碑50=300 → 升3溢出0；p02 6 球 → 60 → 升0溢出60
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            await service.record_match("p02", "2026-08-01", "", {"goal": 6}, "admin")
            await service.advance_period("成长期2", True)
            export = ExportService(service._db, dao)
            rows = await export.rows_for_period(1)
            assert len(rows) == 2
            by_uid = {r["player_uid"]: r for r in rows}
            # p01: 期初 0（首期无结转），期内获得 = 250 数据 + 50 里程碑 = 300，期末 300
            assert by_uid["p01"]["xp_start"] == 0.0
            assert by_uid["p01"]["xp_gained"] == 300.0
            assert by_uid["p01"]["xp_end"] == 300.0
            assert by_uid["p01"]["level_gained"] == 3
            assert by_uid["p01"]["xp_carryover"] == 0.0
            # p02: 期初 0，期内获得 60，期末 60
            assert by_uid["p02"]["xp_start"] == 0.0
            assert by_uid["p02"]["xp_gained"] == 60.0
            assert by_uid["p02"]["xp_end"] == 60.0
            assert by_uid["p02"]["level_gained"] == 0
            assert by_uid["p02"]["xp_carryover"] == 60.0
            # 按期末总经验降序
            assert rows[0]["player_uid"] == "p01"
            # 球队字段随快照 JOIN 带出
            assert by_uid["p01"]["player_team"] == "A队"
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_current_rows():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            export = ExportService(service._db, dao)
            rows = await export.rows_current()
            by_uid = {r["player_uid"]: r for r in rows}
            # 当前期：期初 0（首期无结转），期内获得 300（含里程碑 50），期末 = 球员 xp
            assert by_uid["p01"]["xp_start"] == 0.0
            assert by_uid["p01"]["xp_gained"] == 300.0
            assert by_uid["p01"]["xp_end"] == 300.0
            assert by_uid["p01"]["level_gained"] is None
            assert by_uid["p01"]["xp_carryover"] is None
            assert by_uid["p01"]["level"] == 1
            assert by_uid["p02"]["xp_start"] == 0.0
            assert by_uid["p02"]["xp_gained"] == 0.0
            assert by_uid["p02"]["xp_end"] == 0.0
            # 按当前经验降序
            assert rows[0]["player_uid"] == "p01"
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_overwrite_keeps_identity():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 5}, "admin")
            await service.record_match("p01", "2026-08-01", "", {"goal": 3}, "admin")
            export = ExportService(service._db, dao)
            rows = await export.rows_current()
            r = rows[0]
            # 覆盖录入后：期内获得 = 期末 − 期初（5 球记录被替换为 3 球）
            assert r["xp_gained"] == 30.0
            assert r["xp_start"] == 0.0
            assert r["xp_end"] == 30.0
            await service.advance_period("成长期2", True)
            rows1 = await export.rows_for_period(1)
            r1 = rows1[0]
            assert r1["xp_gained"] == 30.0
            assert r1["xp_start"] == 0.0
            assert r1["xp_end"] == 30.0
            assert r1["level_gained"] == 0
            assert r1["xp_carryover"] == 30.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_includes_period_bonus():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            # 25 球 → 数据 250 + 成长期里程碑 +50 = 300；期内获得必须含奖励
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            export = ExportService(service._db, dao)
            rows = await export.rows_current()
            r = rows[0]
            assert r["xp_start"] == 0.0
            assert r["xp_gained"] == 300.0
            assert r["xp_end"] == 300.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_carryover_strategy_next_start():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            await service.record_match("p02", "2026-08-01", "", {"goal": 6}, "admin")
            # 保留：p02 溢出 60 结转 → 新期期初 = 60
            await service.advance_period("成长期2", True)
            export = ExportService(service._db, dao)
            rows = await export.rows_current()
            p02 = {r["player_uid"]: r for r in rows}["p02"]
            assert p02["xp_start"] == 60.0
            assert p02["xp_gained"] == 0.0
            assert p02["xp_end"] == 60.0
            # 清零：第二期无新增，推进清零 → 新期期初 = 0
            await service.advance_period("成长期3", False)
            rows2 = await export.rows_current()
            p02b = {r["player_uid"]: r for r in rows2}["p02"]
            assert p02b["xp_start"] == 0.0
            assert p02b["xp_end"] == 0.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_build_file_xlsx():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            export = ExportService(service._db, dao)
            rows = await export.rows_current()
            path, fmt = export.build_file(rows, "测试标题", "测试副标题", "当前成长期1_成长数据")
            assert fmt == "xlsx" and path.suffix == ".xlsx" and path.is_file()
            from openpyxl import load_workbook

            wb = load_workbook(path)
            try:
                ws = wb.worksheets[0]
                assert ws.title == "成长数据"
                assert ws["A1"].value == "测试标题"
                assert ws["A2"].value == "测试副标题"
                assert [ws.cell(row=3, column=j).value for j in range(1, 9)] == _HEADERS
                # 数据行：p01 期初0 / 获得300（含里程碑50）/ 期末300 / 未结算 ×2
                row = [ws.cell(row=4, column=j).value for j in range(1, 9)]
                assert row[:3] == ["p01", "球员一", "A队"]
                assert row[3] == 0.0 and row[4] == 300.0 and row[5] == 300.0
                assert row[6] == "未结算" and row[7] == "未结算"
                # 汇总行（2 名球员 → 第 6 行）
                assert ws.cell(row=6, column=1).value == "合计"
                assert ws.cell(row=6, column=6).value == 300.0
                # 冻结窗格 + 标题合并
                assert ws.freeze_panes == "A4"
                merged = [str(rng) for rng in ws.merged_cells.ranges]
                assert "A1:H1" in merged and "A2:H2" in merged
            finally:
                wb.close()
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_build_file_fallback(monkeypatch):
    import astrbot_plugin_whleague_growth_system.services.export_service as es

    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            export = ExportService(service._db, dao)
            rows = await export.rows_current()

            def _boom(*a, **k):
                raise RuntimeError("格式构建失败")

            # openpyxl 不可用 → 降级 CSV（utf-8-sig）
            monkeypatch.setattr(es, "_build_xlsx", _boom)
            path, fmt = export.build_file(rows, "标题", "副标题", "当前成长期1_成长数据")
            assert fmt == "csv" and path.suffix == ".csv"
            assert path.read_bytes()[:3] == b"\xef\xbb\xbf"
            text = path.read_text(encoding="utf-8-sig")
            assert "球员ID" in text and "300" in text and "未结算" in text
            # CSV 也不可用 → 降级 TXT
            monkeypatch.setattr(es, "_build_csv", _boom)
            path2, fmt2 = export.build_file(rows, "标题", "副标题", "当前成长期1_成长数据")
            assert fmt2 == "txt" and path2.suffix == ".txt"
            text2 = path2.read_text(encoding="utf-8")
            assert "合计" in text2 and "|" in text2 and "未结算" in text2
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_export_build_export_current():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            export = ExportService(service._db, dao)
            res = await export.build_export(None)
            assert "当前成长期" in res["title"]
            assert res["path"].is_file()
            assert len(res["rows"]) == 2
            # 不存在的期号 / 无快照期
            with pytest.raises(ValueError):
                await export.build_export(99)
            with pytest.raises(ValueError):
                await export.build_export(2)
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_preview_handler_output():
    import types

    from astrbot_plugin_whleague_growth_system.handlers.player import PlayerHandler

    class _FakeEvent:
        def __init__(self):
            self.results = []

        def plain_result(self, text):
            self.results.append(text)
            return self

    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            await service.record_match("p01", "2026-08-01", "", {"goal": 25}, "admin")
            export = ExportService(service._db, dao)
            plugin = types.SimpleNamespace(
                dao=dao, growth_service=service, import_service=imp, export_service=export
            )
            ph = PlayerHandler(plugin)
            ev = _FakeEvent()
            async for _ in ph.preview(ev):
                pass
            text = ev.results[0]
            assert "当前成长期" in text and "期内累计获得经验 300" in text
            assert "球员一(p01)" in text
            assert "期初 0" in text and "总 300" in text and "Lv1" in text
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


# ─── 多数据项总和里程碑 + 单场达标奖励（v0.5.0）────────────

def _multikey_rule():
    return {
        "stats": {
            "goal": {"name": "进球", "xp": 10},
            "assist": {"name": "助攻", "xp": 5},
            "rating": {"name": "评分", "xp": 1},
        },
        "milestones": [
            {"stats": ["goal", "assist"], "period": "period", "threshold": 10, "xp": 50},
            {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10},
        ],
        "level_xp": 100,
    }


def _saved_rule(rule: dict) -> dict:
    """规范化并保存：save_rule 原样入库，须与生产流程一致先经 normalize。"""
    return parse_rule_json(json.dumps(rule), 100)


def test_parse_rule_multikey_stats_array():
    rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
    m = rule["milestones"][0]
    assert m["stat_keys"] == ["assist", "goal"]  # 排序去重
    assert m["period"] == "period" and m["threshold"] == 10 and m["xp"] == 50


def test_parse_rule_multikey_comma_stat():
    data = _multikey_rule()
    data["milestones"] = [
        {"stat": "goal,assist", "period": "period", "threshold": 10, "xp": 50}
    ]
    rule = parse_rule_json(json.dumps(data), 100)
    assert rule["milestones"][0]["stat_keys"] == ["assist", "goal"]


def test_parse_rule_multikey_conflict():
    data = _multikey_rule()
    data["milestones"] = [
        {"stat": "goal", "stats": ["goal", "assist"], "period": "period",
         "threshold": 10, "xp": 50}
    ]
    with pytest.raises(RuleError):
        parse_rule_json(json.dumps(data), 100)


def test_parse_rule_multikey_unknown_stat():
    data = _multikey_rule()
    data["milestones"] = [
        {"stats": ["goal", "nope"], "period": "period", "threshold": 10, "xp": 50}
    ]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(data), 100)
    assert "nope" in str(e.value)


def test_parse_rule_multikey_bands_stat_rejected():
    rule = _bands_rule()
    rule["milestones"] = [
        {"stats": ["goal", "rating"], "period": "period", "threshold": 10, "xp": 50}
    ]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(rule), 100)
    assert "rating" in str(e.value)


def test_parse_rule_multikey_repeat_rejected():
    data = _multikey_rule()
    data["milestones"] = [
        {"stats": ["goal", "assist"], "period": "period", "step": 10, "xp": 50}
    ]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(data), 100)
    assert "多数据项总和暂不支持" in str(e.value)


def test_parse_rule_multikey_duplicate():
    data = _multikey_rule()
    data["milestones"] = [
        {"stats": ["assist", "goal"], "period": "period", "threshold": 10, "xp": 50},
        {"stats": ["goal", "assist"], "period": "period", "threshold": 10, "xp": 50},
    ]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(data), 100)
    assert "重复定义" in str(e.value)


def test_parse_rule_multikey_dedupe_and_empty():
    # 重复 key 去重后剩单 key → 归一化为普通单数据项里程碑
    data = _multikey_rule()
    data["milestones"] = [
        {"stats": ["goal", "goal"], "period": "period", "threshold": 10, "xp": 50}
    ]
    rule = parse_rule_json(json.dumps(data), 100)
    m = rule["milestones"][0]
    assert "stat_keys" not in m and m["stat"] == "goal"
    # 仅逗号的 stat 报干净错误而非崩溃
    data["milestones"] = [{"stat": ",", "period": "period", "threshold": 10, "xp": 50}]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(data), 100)
    assert "缺少数据项 stat" in str(e.value)


def test_parse_rule_table_multikey_comma():
    rows = [
        ["type", "stat", "name", "xp", "period", "threshold"],
        ["stat", "goal", "进球", "10", "", ""],
        ["stat", "assist", "助攻", "5", "", ""],
        ["milestone", "goal,assist", "", "50", "period", "10"],
    ]
    rule = parse_rule_table(rows, DEFAULT_CONFIG, 100)
    m = [x for x in rule["milestones"] if "stat_keys" in x][0]
    assert m["stat_keys"] == ["assist", "goal"]
    assert m["threshold"] == 10 and m["xp"] == 50


def test_parse_rule_match_ok():
    data = _multikey_rule()
    data["milestones"] = [
        {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10}
    ]
    rule = parse_rule_json(json.dumps(data), 100)
    m = rule["milestones"][0]
    assert m["period"] == "match" and m["stat"] == "rating"
    assert m["threshold"] == 9.0


def test_parse_rule_match_bands_ok():
    rule = _bands_rule()
    rule["milestones"] = [
        {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 15}
    ]
    r = parse_rule_json(json.dumps(rule), 100)
    assert r["milestones"][0]["period"] == "match"


def test_parse_rule_match_repeat_rejected():
    data = _multikey_rule()
    data["milestones"] = [
        {"stat": "rating", "period": "match", "step": 3, "xp": 10}
    ]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(data), 100)
    assert "单场达标" in str(e.value)


def test_parse_rule_match_multikey_rejected():
    data = _multikey_rule()
    data["milestones"] = [
        {"stats": ["goal", "assist"], "period": "match", "threshold": 10, "xp": 50}
    ]
    with pytest.raises(RuleError) as e:
        parse_rule_json(json.dumps(data), 100)
    assert "仅支持单个数据项" in str(e.value)


def test_parse_rule_table_match_period():
    rows = [
        ["type", "stat", "name", "xp", "period", "threshold"],
        ["stat", "rating", "评分", "1", "", ""],
        ["milestone", "rating", "", "10", "match", "9"],
    ]
    rule = parse_rule_table(rows, DEFAULT_CONFIG, 100)
    m = rule["milestones"][0]
    assert m["period"] == "match" and m["threshold"] == 9.0


def test_format_rule_multikey_and_match():
    rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
    text = format_rule(rule)
    assert "助攻+进球" in text and "累计合计达 10" in text
    assert "评分 单场达 9" in text and "额外 10 经验" in text


def test_multikey_milestone_calc_and_idempotent():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            # 移除 match 里程碑，仅测多 key 累计
            rule["milestones"] = [
                {"stats": ["goal", "assist"], "period": "period", "threshold": 10, "xp": 50}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r1 = await service.record_match("p01", "2026-08-01", "", {"goal": 5, "assist": 3}, "admin")
            # 5+3=8 <10，未触发；数据经验 5*10+3*5=65
            assert r1["bonus_xp"] == 0
            r2 = await service.record_match("p01", "2026-08-08", "", {"goal": 2, "assist": 4}, "admin")
            # 累计 14 ≥10 → +50；数据经验 2*10+4*5=40
            assert r2["bonus_xp"] == 50
            assert len(r2["awarded"]) == 1
            p = await dao.get_player("p01")
            assert p["xp"] == 65 + 40 + 50
            r3 = await service.record_match("p01", "2026-08-15", "", {"goal": 1, "assist": 1}, "admin")
            assert r3["bonus_xp"] == 0  # 已颁发，不重复
            p = await dao.get_player("p01")
            assert p["xp"] == 65 + 40 + 50 + 15  # r3 数据经验 1*10+1*5=15
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_multikey_milestone_career():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            rule["milestones"] = [
                {"stats": ["goal", "assist"], "period": "career", "threshold": 50, "xp": 100}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            await service.record_match("p01", "2026-08-01", "", {"goal": 30, "assist": 10}, "admin")
            # 数据经验 30*10+10*5=350；合计 40 <50
            await service.advance_period("成长期2", True)
            r = await service.record_match("p01", "2026-09-01", "", {"goal": 10, "assist": 5}, "admin")
            # 跨期合计 55 ≥50 → 生涯奖励 +100；数据经验 10*10+5*5=125
            assert r["bonus_xp"] == 100
            p = await dao.get_player("p01")
            assert p["xp_total"] == 350 + 125 + 100
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_match_bonus_triggers():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r = await service.record_match("p01", "2026-08-01", "", {"rating": 9.5, "goal": 2}, "admin")
            # 数据经验 9.5*1+2*10=29.5；单场达标 +10
            assert r["match_bonus"] == 10.0
            assert r["bonus_xp"] == 10.0
            assert r["total_xp"] == 39.5
            assert len(r["awarded"]) == 1
            p = await dao.get_player("p01")
            assert p["xp"] == 39.5
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_match_bonus_not_reached():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r = await service.record_match("p01", "2026-08-01", "", {"rating": 8.5}, "admin")
            assert r["match_bonus"] == 0.0
            assert r["awarded"] == []
            p = await dao.get_player("p01")
            assert p["xp"] == 8.5
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_match_bonus_multiple_accumulate():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10},
                {"stat": "rating", "period": "match", "threshold": 9.5, "xp": 20},
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r = await service.record_match("p01", "2026-08-01", "", {"rating": 9.8}, "admin")
            # 两条同时达标 → 全部累加
            assert r["match_bonus"] == 30.0
            assert len(r["awarded"]) == 2
            p = await dao.get_player("p01")
            assert p["xp"] == 9.8 + 30.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_match_bonus_overwrite_recovery():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r1 = await service.record_match("p01", "2026-08-01", "", {"rating": 9.8}, "admin")
            assert r1["xp"] == 19.8
            # 覆盖为不达标：回收旧单场奖励
            r2 = await service.record_match("p01", "2026-08-01", "", {"rating": 8.0}, "admin")
            assert r2["match_bonus"] == 0.0
            assert r2["xp"] == 8.0
            # 再覆盖为达标：补发
            r3 = await service.record_match("p01", "2026-08-01", "", {"rating": 9.7}, "admin")
            assert r3["match_bonus"] == 10.0
            assert r3["xp"] == 19.7
            p = await dao.get_player("p01")
            assert p["xp"] == 19.7
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_match_bonus_overwrite_after_rule_change():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = parse_rule_json(json.dumps(_multikey_rule()), 100)
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 10}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r1 = await service.record_match("p01", "2026-08-01", "", {"rating": 9.5}, "admin")
            assert r1["xp"] == 19.5  # 含旧规则单场奖励 +10
            # 改规则：阈值 9.0 → 9.8，奖励 10 → 20
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.8, "xp": 20}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r2 = await service.record_match("p01", "2026-08-01", "", {"rating": 9.6}, "admin")
            # 旧奖励按存储值回收（10），新规则 9.6<9.8 不触发 → xp=9.6
            assert r2["match_bonus"] == 0.0
            assert r2["xp"] == 9.6
            p = await dao.get_player("p01")
            assert p["xp"] == 9.6
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())


def test_match_bonus_bands_stat():
    service, dao, imp, tmp, env = _make_env()
    try:
        async def _run():
            await _setup(service, dao)
            rule = _bands_rule()
            rule["milestones"] = [
                {"stat": "rating", "period": "match", "threshold": 9.0, "xp": 15}
            ]
            await service.save_rule(_saved_rule(rule), "test", "admin")
            r = await service.record_match("p01", "2026-08-01", "", {"rating": 9.5}, "admin")
            # bands [8~)+20 + 单场达标 15 = 35
            assert r["stat_xp"] == 20.0
            assert r["match_bonus"] == 15.0
            assert r["xp"] == 35.0
            r2 = await service.record_match("p01", "2026-08-08", "", {"rating": 6.5}, "admin")
            # bands [6~8)+10，未达标
            assert r2["stat_xp"] == 10.0
            assert r2["match_bonus"] == 0.0
            assert r2["xp"] == 45.0
        _run_async(_run())
    finally:
        asyncio.run(env["db"].close())
