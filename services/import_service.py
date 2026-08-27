"""文件导入服务：规则 / 球员库 / 比赛数据（JSON / CSV / Excel）。

- xlsx 使用 openpyxl 流式读取（read_only）；csv 自动探测编码（utf-8-sig → gbk → utf-8）
- 文件名前缀决定类型：规则_（规则）、球员_（球员库）、比赛_（比赛数据）
- 群内发文件 → 捕获 → 预览 → /成长 导入 确认 <文件> [类型] 执行
- imports 目录文件数超上限时自动删除最旧
"""

import csv
import json
import os
from pathlib import Path

from astrbot.api import logger

from ..utils.security import sanitize_filename, sanitize_text, sanitize_uid
from . import rule_parser

_ALLOWED_EXTS = (".xlsx", ".csv")
_RULE_EXTS = (".json", ".xlsx", ".csv")

_PREFIX_KIND = (
    ("规则", "rule"),
    ("球员", "players"),
    ("比赛", "matches"),
)

_KIND_NAMES = {"rule": "规则", "players": "球员", "matches": "比赛"}


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return format(value, ".15f").rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_csv(file_path: Path, max_rows: int):
    raw = file_path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    rows = []
    for i, line in enumerate(csv.reader(text.splitlines())):
        if i > max_rows:
            break
        rows.append(line)
    return rows


def _parse_xlsx(file_path: Path, max_rows: int):
    from openpyxl import load_workbook

    rows = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > max_rows:
                break
            rows.append([_cell_to_str(v) for v in row])
    finally:
        wb.close()
    return rows


def _parse_json(file_path: Path):
    return json.loads(file_path.read_text(encoding="utf-8-sig"))


def kind_from_name(file_name: str) -> str | None:
    """按文件名前缀推断类型：规则_/球员_/比赛_ → rule/players/matches。"""
    for prefix, kind in _PREFIX_KIND:
        if file_name.startswith(prefix):
            return kind
    return None


class GrowthImportService:
    def __init__(self, db, dao, cfg_get, growth_service):
        self._db = db
        self._dao = dao
        self._cfg_get = cfg_get
        self._growth = growth_service

    @property
    def imports_dir(self) -> Path:
        base = Path(self._db.db_path).parent
        d = base / "imports"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _col(self, key: str) -> int:
        return int(self._cfg_get(key, 0) or 0)

    def kind_from_name(self, file_name: str) -> str | None:
        """实例方法：委托模块级 kind_from_name（群文件捕获/命令共用）。"""
        return kind_from_name(file_name)

    def list_files(self) -> list:
        files = [
            p for p in self.imports_dir.iterdir()
            if p.is_file() and p.suffix.lower() in _ALLOWED_EXTS
        ]
        files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return files

    def check_file(self, file_name: str, kind: str | None = None) -> Path:
        safe = sanitize_filename(file_name)
        p = self.imports_dir / safe
        if not p.is_file() or p.parent.resolve() != self.imports_dir.resolve():
            raise FileNotFoundError(
                f"文件「{file_name}」不存在于导入目录。请先在群内发送该文件，或检查文件名"
            )
        ext = p.suffix.lower()
        if kind == "rule":
            if ext not in _RULE_EXTS:
                raise ValueError("规则文件仅支持 .json / .xlsx / .csv")
        elif ext not in _ALLOWED_EXTS:
            raise ValueError("仅支持 .xlsx / .csv 文件")
        size_mb = int(self._cfg_get("import_max_file_size_mb", 50) or 50)
        if p.stat().st_size > size_mb * 1024 * 1024:
            raise ValueError(f"文件超过大小上限（{size_mb} MB）")
        return p

    def cleanup_oldest(self) -> int:
        max_files = int(self._cfg_get("import_max_files", 50) or 50)
        files = self.list_files()
        removed = 0
        while len(files) > max_files:
            old = files.pop()
            try:
                old.unlink()
                removed += 1
            except OSError as e:
                logger.warning(f"Failed to remove import file {old}: {e}")
        return removed

    def save_uploaded(self, file_path: str, file_name: str) -> Path:
        safe_name = sanitize_filename(file_name)
        ext = Path(safe_name).suffix.lower()
        if ext not in _ALLOWED_EXTS and ext not in _RULE_EXTS:
            raise ValueError("仅支持 .json / .xlsx / .csv 文件")
        size_mb = int(self._cfg_get("import_max_file_size_mb", 50) or 50)
        size = os.path.getsize(file_path)
        if size > size_mb * 1024 * 1024:
            raise ValueError(f"文件超过大小上限（{size_mb} MB）")
        target = self.imports_dir / safe_name
        target.write_bytes(Path(file_path).read_bytes())
        self.cleanup_oldest()
        return target

    def _read_rows(self, file_path) -> list:
        file_path = Path(file_path)
        max_rows = int(self._cfg_get("import_max_rows", 50000) or 50000)
        if file_path.suffix.lower() == ".xlsx":
            return _parse_xlsx(file_path, max_rows)
        return _parse_csv(file_path, max_rows)

    # ─── 规则解析 ──────────────────────────────────────────

    def parse_rule_file(self, file_path: Path) -> dict:
        """解析规则文件为规范结构，失败抛 rule_parser.RuleError。"""
        default_level_xp = round(float(self._cfg_get("default_level_xp", 100) or 100), 1)
        ext = file_path.suffix.lower()
        if ext == ".json":
            data = _parse_json(file_path)
            return rule_parser.normalize_rule(data, default_level_xp)
        rows = self._read_rows(file_path)
        return rule_parser.parse_rule_table(rows, self._cfg_get, default_level_xp)

    # ─── 球员库解析 ────────────────────────────────────────

    def parse_players_file(self, file_path: Path) -> tuple[list, list, int]:
        """解析球员库文件，返回 (数据行 [(uid, name, team)], 错误, 跳过行数)。"""
        rows = self._read_rows(file_path)
        col_uid = self._col("import_col_uid")
        col_name = self._col("import_col_name_player")
        col_team = self._col("import_col_team")

        data: list = []
        errors: list = []
        skipped = 0
        for idx, row in enumerate(rows, start=1):
            if not row or not any(str(c).strip() for c in row):
                skipped += 1
                continue
            uid = sanitize_uid(row[col_uid - 1]) if 0 < col_uid <= len(row) else ""
            name = sanitize_text(row[col_name - 1]) if 0 < col_name <= len(row) else ""
            team = sanitize_text(row[col_team - 1]) if 0 < col_team <= len(row) else ""
            if not uid:
                errors.append(f"第{idx}行: 球员 ID 为空或非法")
                continue
            data.append((uid, name, team))
        return data, errors, skipped

    # ─── 比赛数据解析 ──────────────────────────────────────

    def parse_matches_file(self, file_path: Path, rule: dict) -> tuple[list, list, int]:
        """解析比赛数据文件，返回 (录入项列表, 错误, 跳过行数)。

        首行若在日期/球员列位上有值且其余列有匹配规则数据项（键或显示名），视为表头；
        数据行解析后无任何有效数据列则记入错误，不生成 0 经验记录。
        球员列可填球员 ID 或球员姓名（导入时自动匹配，见 growth_service._resolve_player）。
        """
        rows = self._read_rows(file_path)
        col_date = self._col("import_col_match_date")
        col_uid = self._col("import_col_match_uid")
        # 表头单元格 → 数据项键：兼容英文键与中文显示名
        header_to_key = {}
        for key, meta in rule["stats"].items():
            header_to_key[key] = key
            if meta.get("name"):
                header_to_key[meta["name"]] = key

        entries: list = []
        errors: list = []
        skipped = 0
        header = None
        for idx, row in enumerate(rows, start=1):
            if not row or not any(str(c).strip() for c in row):
                skipped += 1
                continue
            if header is None:
                candidate = [_cell_to_str(c).strip() for c in row]
                # 仅当首行其余列（非日期/球员列位）有匹配数据项时才视为表头
                matched = any(
                    i + 1 not in (col_date, col_uid) and h in header_to_key
                    for i, h in enumerate(candidate)
                )
                if matched:
                    header = candidate
                    continue
                # 首行是数据行，不消费
                header = []
            date_raw = _cell_to_str(row[col_date - 1]) if 0 < col_date <= len(row) else ""
            uid_raw = _cell_to_str(row[col_uid - 1]) if 0 < col_uid <= len(row) else ""
            uid = sanitize_uid(uid_raw)
            if not uid:
                errors.append(f"第{idx}行: 球员 ID/姓名 为空或非法")
                continue
            try:
                from ..utils.security import parse_date
                match_date = parse_date(date_raw)
            except ValueError as e:
                errors.append(f"第{idx}行: {e}")
                continue
            stats = {}
            for i, cell in enumerate(row):
                if i >= len(header):
                    break
                key = header_to_key.get(header[i])
                if key is None:
                    continue
                v = _cell_to_str(cell)
                if not v:
                    continue
                try:
                    from ..utils.security import parse_num
                    stats[key] = parse_num(v)
                except ValueError:
                    errors.append(f"第{idx}行: 数据项 {key} 值非法: {v}")
                    continue
            if not stats:
                errors.append(f"第{idx}行: 无任何有效数据列，已跳过（请核对表头与规则数据项是否一致）")
                continue
            entries.append(
                {"player_uid": uid, "match_date": match_date, "opponent": "", "stats": stats}
            )
        return entries, errors, skipped

    # ─── 预览与确认 ────────────────────────────────────────

    async def preview(self, file_path: Path, kind: str) -> str:
        """生成导入预览文本。"""
        if kind == "rule":
            rule = self.parse_rule_file(file_path)
            return rule_parser.format_rule(rule)
        if kind == "players":
            data, errors, skipped = self.parse_players_file(file_path)
            lines = [f"可导入 {len(data)} 名球员（跳过 {skipped} 行）"]
            for uid, name, team in data[:5]:
                lines.append(f"· {uid} {name} {team}".rstrip())
            if errors:
                lines.append(f"⚠️ {len(errors)} 行错误: {errors[0]}")
            return "\n".join(lines)
        if kind == "matches":
            return await self._preview_matches(file_path)
        raise ValueError(f"未知导入类型: {kind}")

    async def _preview_matches(self, file_path: Path) -> str:
        rule = await self._growth.get_rule()
        if rule is None:
            raise ValueError("尚未导入成长规则，无法解析比赛文件")
        entries, errors, skipped = self.parse_matches_file(file_path, rule)
        lines = [f"可导入 {len(entries)} 条球员记录（跳过 {skipped} 行）"]
        for e in entries[:5]:
            stats = " ".join(f"{k}={v:g}" for k, v in e["stats"].items())
            lines.append(f"· {e['match_date']} {e['player_uid']} {stats}".rstrip())
        if errors:
            lines.append(f"⚠️ {len(errors)} 行错误: {errors[0]}")
        return "\n".join(lines)

    # ─── 确认执行 ──────────────────────────────────────────

    async def confirm_rule_import(self, file_name: str, kind: str, created_by: str) -> dict:
        file_path = self.check_file(file_name, kind)
        rule = self.parse_rule_file(file_path)
        await self._growth.save_rule(rule, file_name, created_by)
        return {"rule": rule}

    async def confirm_players_import(self, file_name: str, kind: str, created_by: str) -> dict:
        file_path = self.check_file(file_name, kind)
        data, errors, _ = self.parse_players_file(file_path)
        if not data:
            raise ValueError("没有可导入的球员数据")

        batch_size = max(1, int(self._cfg_get("import_batch_size", 5000) or 5000))
        added = 0
        updated = 0

        def _chunks():
            for i in range(0, len(data), batch_size):
                yield data[i : i + batch_size]

        for chunk in _chunks():
            async def _tx(conn):
                add = 0
                upd = 0
                for uid, name, team in chunk:
                    existing = await self._dao.get_player_conn(conn, uid)
                    await self._dao.upsert_player(conn, uid, name, team, created_by)
                    if existing:
                        upd += 1
                    else:
                        add += 1
                return add, upd

            a, u = await self._db.execute_transaction(_tx)
            added += a
            updated += u
        return {"added": added, "updated": updated, "errors": errors}

    async def confirm_matches_import(self, file_name: str, kind: str, created_by: str) -> dict:
        file_path = self.check_file(file_name, kind)
        rule = await self._growth.get_rule()
        if rule is None:
            raise ValueError("尚未导入成长规则，无法导入比赛数据")
        entries, errors, skipped = self.parse_matches_file(file_path, rule)
        if not entries:
            raise ValueError("没有可导入的比赛数据")
        result = await self._growth.record_match_batch(entries, created_by)
        return {"ok": result["ok"], "errors": errors, "skipped": skipped}

    async def import_pending(self, pending_id: int, created_by: str) -> dict:
        """执行一条待确认导入，成功后标记完成。"""
        pending = await self._dao.get_pending(pending_id)
        if pending is None:
            raise ValueError("待确认导入不存在")
        if pending["status"] != "pending":
            raise ValueError(f"该导入已处理（状态: {pending['status']}）")
        kind = pending["kind"]
        file_name = pending["file_name"]
        if kind == "rule":
            result = await self.confirm_rule_import(file_name, kind, created_by)
        elif kind == "players":
            result = await self.confirm_players_import(file_name, kind, created_by)
        elif kind == "matches":
            result = await self.confirm_matches_import(file_name, kind, created_by)
        else:
            raise ValueError(f"未知导入类型: {kind}")
        await self._dao.update_pending_status(pending_id, "done")
        result["kind"] = kind
        result["file_name"] = file_name
        return result
