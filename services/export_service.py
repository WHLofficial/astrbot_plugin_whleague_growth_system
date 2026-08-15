"""成长数据导出服务：历史成长期 / 当前成长期 → Excel（优先）→ CSV → TXT 降级生成。

- Excel 使用 openpyxl 做标题 + 元信息 + 表头 + 斑马纹等美化排版；生成失败依次降级 CSV、TXT
- 导出文件写入插件数据目录下 exports/，由调用方决定发群或附服务器路径
- 行字段：期初经验（= 上一期结转，首期 0）、期内获得经验（= 期末总经验 − 期初，含里程碑/重复奖励）、期末总经验、升级数、升级后剩余经验
"""

import csv
import unicodedata
from datetime import datetime
from pathlib import Path

from astrbot.api import logger

from ..utils.security import fmt_xp, sanitize_filename

_HEADERS = [
    "球员ID",
    "姓名",
    "球队",
    "期初经验",
    "期内获得经验",
    "期末总经验",
    "升级数",
    "升级后剩余经验",
]


def _disp_width(text: str) -> int:
    """显示宽度：CJK 全角字符按 2 计，用于列宽估算与等宽对齐。"""
    return sum(
        2 if unicodedata.east_asian_width(ch) in ("F", "W") else 1
        for ch in str(text)
    )


def _level_text(value) -> str:
    return "未结算" if value is None else str(int(value))


def _carry_text(value) -> str:
    return "未结算" if value is None else fmt_xp(value)


def _pad(text: str, width: int, align: str = "left") -> str:
    pad = max(0, width - _disp_width(text))
    return text + " " * pad if align == "left" else " " * pad + text


def _summary(rows: list) -> dict:
    start = round(sum(float(r["xp_start"]) for r in rows), 1)
    gained = round(sum(float(r["xp_gained"]) for r in rows), 1)
    end = round(sum(float(r["xp_end"]) for r in rows), 1)
    upgraded = sum(
        1 for r in rows if r["level_gained"] is not None and r["level_gained"] > 0
    )
    settled = any(r["level_gained"] is not None for r in rows)
    upgraded_text = str(upgraded) if settled else "—"
    return {
        "start": start,
        "gained": gained,
        "end": end,
        "upgraded_text": upgraded_text,
    }


# ─── 各格式构建（内部导入，失败逐级降级）──────────────────


def _build_xlsx(
    path: Path, title: str, subtitle: str, headers: list, rows: list, summary: dict
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "成长数据"
    ncols = len(headers)
    last_col = get_column_letter(ncols)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra = PatternFill("solid", fgColor="F2F7FB")

    # 标题行（合并、大字白字深底）
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 元信息行（期号/期名/生成时间）
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = subtitle
    c.font = Font(size=10, color="404040")
    c.fill = PatternFill("solid", fgColor="E7E6E6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # 表头
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[3].height = 20

    # 数据行（斑马纹，升级数列语义着色）
    for i, row in enumerate(rows):
        r = 4 + i
        values = [
            row["player_uid"],
            row["player_name"],
            row["player_team"],
            float(row["xp_start"]),
            float(row["xp_gained"]),
            float(row["xp_end"]),
            row["level_gained"],
            row["xp_carryover"],
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = border
            c.alignment = Alignment(
                horizontal="left" if j in (2, 3) else "center", vertical="center"
            )
            if i % 2 == 1:
                c.fill = zebra
        lv = row["level_gained"]
        lv_cell = ws.cell(row=r, column=7)
        if lv is None:
            lv_cell.value = "未结算"
            lv_cell.font = Font(color="808080")
        elif lv > 0:
            lv_cell.font = Font(bold=True, color="006100")
        else:
            lv_cell.font = Font(color="808080")
        co_cell = ws.cell(row=r, column=8)
        if row["xp_carryover"] is None:
            co_cell.value = "未结算"
            co_cell.font = Font(color="808080")

    # 汇总行
    sr = 4 + len(rows)
    ws.merge_cells(f"A{sr}:C{sr}")
    c = ws.cell(row=sr, column=1, value="合计")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    sum_values = [summary["start"], summary["gained"], summary["end"], summary["upgraded_text"]]
    for j, v in zip((4, 5, 6, 7), sum_values):
        c = ws.cell(row=sr, column=j, value=v)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    c = ws.cell(row=sr, column=1)
    c.border = border
    ws.cell(row=sr, column=2).border = border
    ws.cell(row=sr, column=3).border = border
    ws.cell(row=sr, column=8).border = border

    # 列宽（中文按 2 字符宽估算）+ 冻结标题与表头
    widths = [_disp_width(h) for h in headers]
    for row in rows:
        values = [
            str(row["player_uid"]),
            str(row["player_name"]),
            str(row["player_team"]),
            fmt_xp(row["xp_start"]),
            fmt_xp(row["xp_gained"]),
            fmt_xp(row["xp_end"]),
            _level_text(row["level_gained"]),
            _carry_text(row["xp_carryover"]),
        ]
        for j, v in enumerate(values):
            widths[j] = max(widths[j], _disp_width(v))
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 8), 26)
    ws.freeze_panes = "A4"

    wb.save(str(path))


def _build_csv(
    path: Path, title: str, subtitle: str, headers: list, rows: list, summary: dict
) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([title])
        w.writerow([subtitle])
        w.writerow(headers)
        for row in rows:
            w.writerow(
                [
                    row["player_uid"],
                    row["player_name"],
                    row["player_team"],
                    fmt_xp(row["xp_start"]),
                    fmt_xp(row["xp_gained"]),
                    fmt_xp(row["xp_end"]),
                    _level_text(row["level_gained"]),
                    _carry_text(row["xp_carryover"]),
                ]
            )
        w.writerow(
            [
                "合计",
                "",
                "",
                fmt_xp(summary["start"]),
                fmt_xp(summary["gained"]),
                fmt_xp(summary["end"]),
                summary["upgraded_text"],
                "",
            ]
        )


def _build_txt(
    path: Path, title: str, subtitle: str, headers: list, rows: list, summary: dict
) -> None:
    def to_row(row) -> list:
        return [
            row["player_uid"],
            row["player_name"],
            row["player_team"],
            fmt_xp(row["xp_start"]),
            fmt_xp(row["xp_gained"]),
            fmt_xp(row["xp_end"]),
            _level_text(row["level_gained"]),
            _carry_text(row["xp_carryover"]),
        ]

    data_rows = [to_row(r) for r in rows]
    sum_row = [
        "合计",
        "",
        "",
        fmt_xp(summary["start"]),
        fmt_xp(summary["gained"]),
        fmt_xp(summary["end"]),
        summary["upgraded_text"],
        "",
    ]
    widths = [_disp_width(h) for h in headers]
    for r in data_rows + [sum_row]:
        for j, v in enumerate(r):
            widths[j] = max(widths[j], _disp_width(v))
    widths = [min(max(w + 2, 6), 24) for w in widths]

    def line(cells, numeric: bool) -> str:
        parts = []
        for j, cell in enumerate(cells):
            align = "right" if numeric and j >= 3 else "left"
            parts.append(_pad(str(cell), widths[j], align))
        return "| " + " | ".join(parts) + " |"

    sep = "+" + "+".join("-" * w for w in widths) + "+"
    lines = [title, subtitle, sep, line(headers, numeric=False), sep]
    for r in data_rows:
        lines.append(line(r, numeric=True))
    lines.append(sep)
    lines.append(line(sum_row, numeric=True))
    lines.append(sep)
    path.write_text("\n".join(lines), encoding="utf-8")


class ExportService:
    def __init__(self, db, dao):
        self._db = db
        self._dao = dao

    @property
    def exports_dir(self) -> Path:
        base = Path(self._db.db_path).parent
        d = base / "exports"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _now(self) -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    async def _start_xp_map(self, period_no: int) -> dict:
        """该成长期开始时每位球员的账面经验：上一期快照的溢出结转；首期为 0。"""
        if period_no <= 1:
            return {}
        return await self._dao.get_snapshots_carryover(period_no - 1)

    # ─── 行数据 ────────────────────────────────────────────

    async def rows_for_period(self, period_no: int) -> list:
        """历史成长期行：期末总经验来自快照，期初 = 上一期结转，期内获得 = 期末 − 期初。"""
        snapshots = await self._dao.list_period_snapshots(period_no)
        start_map = await self._start_xp_map(period_no)
        rows = []
        for s in snapshots:
            xp_period = round(float(s["xp_period"]), 1)
            xp_start = start_map.get(s["player_uid"], 0.0)
            rows.append(
                {
                    "player_uid": s["player_uid"],
                    "player_name": s["player_name"],
                    "player_team": s["player_team"] or "",
                    "xp_start": xp_start,
                    "xp_gained": round(xp_period - xp_start, 1),
                    "xp_end": xp_period,
                    "level_gained": int(s["level_gained"]),
                    "xp_carryover": round(float(s["xp_carryover"]), 1),
                }
            )
        return rows

    async def rows_current(self) -> list:
        """当前未结束成长期行：期末总经验为球员当前 xp，升级数/结转未结算。"""
        period = await self._dao.get_current_period()
        if period is None:
            return []
        period_no = period["period_no"]
        start_map = await self._start_xp_map(period_no)
        players = await self._dao.list_all_active_players()
        rows = []
        for p in players:
            xp = round(float(p["xp"]), 1)
            xp_start = start_map.get(p["player_uid"], 0.0)
            rows.append(
                {
                    "player_uid": p["player_uid"],
                    "player_name": p["name"],
                    "player_team": p["team"] or "",
                    "xp_start": xp_start,
                    "xp_gained": round(xp - xp_start, 1),
                    "xp_end": xp,
                    "level_gained": None,
                    "xp_carryover": None,
                    "level": int(p["level"]),
                }
            )
        rows.sort(key=lambda r: r["xp_end"], reverse=True)
        return rows

    # ─── 文件生成 ──────────────────────────────────────────

    def build_file(
        self, rows: list, title: str, subtitle: str, base_name: str
    ) -> tuple:
        """按 Excel → CSV → TXT 顺序降级生成，返回 (path, 格式)。"""
        safe = sanitize_filename(base_name)
        headers = _HEADERS
        summary = _summary(rows)
        fmt = "xlsx"
        path = self.exports_dir / f"{safe}.xlsx"
        try:
            _build_xlsx(path, title, subtitle, headers, rows, summary)
        except Exception as e:
            logger.warning(f"Excel 导出失败，降级 CSV: {e}")
            fmt = "csv"
            path = path.with_suffix(".csv")
            try:
                _build_csv(path, title, subtitle, headers, rows, summary)
            except Exception as e2:
                logger.warning(f"CSV 导出失败，降级 TXT: {e2}")
                fmt = "txt"
                path = path.with_suffix(".txt")
                _build_txt(path, title, subtitle, headers, rows, summary)
        return path, fmt

    async def build_export(self, period_no: int | None) -> dict:
        """组装导出：period_no 为 None 时导出当前未结束成长期。"""
        if period_no is None:
            period = await self._dao.get_current_period()
            if period is None:
                raise ValueError("不存在当前成长期，无法导出")
            rows = await self.rows_current()
            title = f"当前成长期 #{period['period_no']} {period['name']} 成长数据"
            base_name = f"当前成长期{period['period_no']}_成长数据"
        else:
            period = await self._db.fetchone(
                "SELECT * FROM growth_periods WHERE period_no=?", (period_no,)
            )
            if period is None:
                raise ValueError(f"成长期 #{period_no} 不存在")
            rows = await self.rows_for_period(period_no)
            if not rows:
                raise ValueError(f"成长期 #{period_no} 尚无快照数据（该期可能从未推进过）")
            title = f"成长期#{period['period_no']} {period['name']} 成长数据"
            base_name = f"成长期{period['period_no']}_成长数据"
        subtitle = f"期号 #{period['period_no']} · {period['name']} · 生成时间 {self._now()}"
        path, fmt = self.build_file(rows, title, subtitle, base_name)
        return {
            "path": path,
            "fmt": fmt,
            "title": title,
            "rows": rows,
            "period": period,
        }
